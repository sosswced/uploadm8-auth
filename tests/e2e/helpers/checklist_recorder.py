"""Append-only checklist log + Excel export for overnight full-app runs."""

from __future__ import annotations

import json
import threading
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Literal

Status = Literal["PASS", "FAIL", "SKIP", "WARN"]

ROOT = Path(__file__).resolve().parents[3]
DEFAULT_LOG = ROOT / "tests" / "e2e" / "artifacts" / "checklist_items.jsonl"

_lock = threading.Lock()


@dataclass
class CheckItem:
    check_id: str
    category: str
    name: str
    status: Status
    detail: str = ""
    duration_ms: int = 0
    ts: str = field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class ChecklistRecorder:
    def __init__(self, path: Path | None = None):
        self.path = path or DEFAULT_LOG
        self.path.parent.mkdir(parents=True, exist_ok=True)

    def record(
        self,
        *,
        check_id: str,
        category: str,
        name: str,
        status: Status,
        detail: str = "",
        duration_ms: int = 0,
    ) -> None:
        item = CheckItem(
            check_id=check_id,
            category=category,
            name=name,
            status=status,
            detail=detail[:2000],
            duration_ms=duration_ms,
        )
        line = json.dumps(asdict(item), ensure_ascii=False)
        with _lock:
            with self.path.open("a", encoding="utf-8") as f:
                f.write(line + "\n")

    def load_all(self) -> list[dict]:
        if not self.path.is_file():
            return []
        out: list[dict] = []
        for line in self.path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line:
                out.append(json.loads(line))
        return out

    def reset(self) -> None:
        if self.path.is_file():
            self.path.unlink()


def export_checklist_excel(items: list[dict], xlsx_path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Full App Checklist"

    headers = [
        "#",
        "Check ID",
        "Category",
        "Name",
        "Status",
        "Duration (ms)",
        "Detail",
        "Timestamp",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_fills = {
        "PASS": PatternFill("solid", fgColor="DCFCE7"),
        "FAIL": PatternFill("solid", fgColor="FEE2E2"),
        "SKIP": PatternFill("solid", fgColor="F3F4F6"),
        "WARN": PatternFill("solid", fgColor="FEF9C3"),
    }

    for i, row in enumerate(items, start=1):
        ws.append(
            [
                i,
                row.get("check_id", ""),
                row.get("category", ""),
                row.get("name", ""),
                row.get("status", ""),
                row.get("duration_ms", 0),
                row.get("detail", ""),
                row.get("ts", ""),
            ]
        )
        st = str(row.get("status", ""))
        fill = status_fills.get(st)
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i + 1, column=col).fill = fill

    # Summary sheet
    summary = wb.create_sheet("Summary")
    total = len(items)
    passed = sum(1 for r in items if r.get("status") == "PASS")
    failed = sum(1 for r in items if r.get("status") == "FAIL")
    skipped = sum(1 for r in items if r.get("status") == "SKIP")
    warned = sum(1 for r in items if r.get("status") == "WARN")
    summary.append(["Metric", "Value"])
    summary.append(["Total checks", total])
    summary.append(["Passed", passed])
    summary.append(["Failed", failed])
    summary.append(["Skipped", skipped])
    summary.append(["Warnings", warned])
    summary.append(["Pass rate", f"{(100.0 * passed / total):.1f}%" if total else "—"])
    summary.append(["Generated (UTC)", datetime.now(timezone.utc).isoformat()])

    by_cat: dict[str, dict[str, int]] = {}
    for r in items:
        cat = r.get("category", "Other")
        st = r.get("status", "FAIL")
        by_cat.setdefault(cat, {"PASS": 0, "FAIL": 0, "SKIP": 0, "WARN": 0})
        by_cat[cat][st] = by_cat[cat].get(st, 0) + 1
    summary.append([])
    summary.append(["Category", "PASS", "FAIL", "SKIP", "WARN"])
    for cat in sorted(by_cat):
        c = by_cat[cat]
        summary.append([cat, c.get("PASS", 0), c.get("FAIL", 0), c.get("SKIP", 0), c.get("WARN", 0)])

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 48
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 60
    ws.column_dimensions["H"].width = 24

    wb.save(xlsx_path)
    return xlsx_path


ARTIFACTS_DIR = ROOT / "tests" / "e2e" / "artifacts"
CHECKLIST_GLOBS = (
    "UploadM8_Full_App_Checklist_*.xlsx",
    "UploadM8_Human_App_Tour_*.xlsx",
)


def find_latest_checklist_xlsx(
    artifacts_dir: Path | None = None,
    *,
    prefer: str | None = None,
) -> Path | None:
    """Return newest checklist Excel under tests/e2e/artifacts."""
    base = artifacts_dir or ARTIFACTS_DIR
    if not base.is_dir():
        return None
    if prefer:
        p = Path(prefer)
        if p.is_file():
            return p.resolve()
        candidate = base / prefer
        if candidate.is_file():
            return candidate.resolve()
    matches: list[Path] = []
    for pattern in CHECKLIST_GLOBS:
        matches.extend(base.glob(pattern))
    if not matches:
        return None
    return max(matches, key=lambda p: p.stat().st_mtime)


def summarize_checklist_excel(xlsx_path: Path) -> dict:
    """Parse checklist Excel for overnight / review summaries."""
    from openpyxl import load_workbook

    path = Path(xlsx_path).resolve()
    if not path.is_file():
        raise FileNotFoundError(path)

    wb = load_workbook(path, read_only=True, data_only=True)
    summary_metrics: dict[str, str | int | float] = {}
    if "Summary" in wb.sheetnames:
        ws_sum = wb["Summary"]
        for row in ws_sum.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            if key in ("Category", ""):
                break
            summary_metrics[key] = row[1]

    ws = wb[wb.sheetnames[0]]
    by_status: dict[str, int] = {"PASS": 0, "FAIL": 0, "SKIP": 0, "WARN": 0}
    by_category: dict[str, dict[str, int]] = {}
    failures: list[dict] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5 or not row[1]:
            continue
        check_id = str(row[1])
        category = str(row[2] or "Other")
        name = str(row[3] or "")
        status = str(row[4] or "FAIL")
        detail = str(row[6] or "") if len(row) > 6 else ""
        by_status[status] = by_status.get(status, 0) + 1
        by_category.setdefault(category, {"PASS": 0, "FAIL": 0, "SKIP": 0, "WARN": 0})
        by_category[category][status] = by_category[category].get(status, 0) + 1
        if status == "FAIL":
            failures.append(
                {
                    "check_id": check_id,
                    "category": category,
                    "name": name,
                    "detail": detail[:500],
                }
            )
    wb.close()

    total = sum(by_status.values())
    passed = by_status.get("PASS", 0)
    return {
        "path": str(path),
        "filename": path.name,
        "total": total,
        "passed": passed,
        "failed": by_status.get("FAIL", 0),
        "skipped": by_status.get("SKIP", 0),
        "warnings": by_status.get("WARN", 0),
        "pass_rate": round(100.0 * passed / total, 1) if total else 0.0,
        "summary_sheet": summary_metrics,
        "by_category": {
            cat: counts for cat, counts in sorted(by_category.items())
        },
        "failures_by_category": {
            cat: sum(1 for f in failures if f["category"] == cat)
            for cat in sorted({f["category"] for f in failures})
        },
        "failures": failures,
    }


def format_checklist_review_markdown(summary: dict, *, max_failures: int = 20) -> str:
    """Human-readable block for /overnight completion summaries."""
    lines = [
        f"### Full App Checklist — `{summary['filename']}`",
        "",
        "| Metric | Value |",
        "|--------|-------|",
        f"| Total | {summary['total']} |",
        f"| Passed | {summary['passed']} |",
        f"| Failed | {summary['failed']} |",
        f"| Skipped | {summary['skipped']} |",
        f"| Warnings | {summary['warnings']} |",
        f"| Pass rate | {summary['pass_rate']}% |",
        "",
    ]
    if summary.get("failures_by_category"):
        lines.append("**Failures by category:**")
        for cat, count in sorted(
            summary["failures_by_category"].items(), key=lambda x: -x[1]
        ):
            lines.append(f"- {cat}: {count}")
        lines.append("")
    fails = summary.get("failures") or []
    if fails:
        lines.append(f"**Top failures** (showing up to {max_failures} of {len(fails)}):")
        for f in fails[:max_failures]:
            detail = f["detail"][:100].replace("\n", " ") if f["detail"] else ""
            suffix = f" — {detail}" if detail else ""
            lines.append(f"- `[{f['category']}]` **{f['check_id']}**: {f['name']}{suffix}")
    lines.append("")
    lines.append(f"Excel: `{summary['path']}`")
    return "\n".join(lines)


if __name__ == "__main__":
    import argparse
    import sys

    parser = argparse.ArgumentParser(description="Summarize UploadM8 checklist Excel for reviews")
    parser.add_argument(
        "xlsx",
        nargs="?",
        help="Path to UploadM8_Full_App_Checklist_*.xlsx (default: newest in artifacts/)",
    )
    parser.add_argument("--json", action="store_true", help="Emit JSON instead of markdown")
    parser.add_argument("--max-failures", type=int, default=25)
    args = parser.parse_args()

    path = find_latest_checklist_xlsx(prefer=args.xlsx)
    if not path:
        print("No checklist Excel found under tests/e2e/artifacts/", file=sys.stderr)
        raise SystemExit(1)
    summary = summarize_checklist_excel(path)
    if args.json:
        print(json.dumps(summary, indent=2, ensure_ascii=False))
    else:
        print(format_checklist_review_markdown(summary, max_failures=args.max_failures))
